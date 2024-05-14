import openpyxl
wb = openpyxl.Workbook()    # wb -> 워크북 인스턴스 객체 생성
sheet = wb.active   # 활성화된 워크시트 객체 가져옴
sheet.title = '회원정보'
header_titles = ['이름', '전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row = 1, column= idx + 1, value = title)
members = [('gildong', '010-1111-2222'), ('chanho', '010-5555-3333'), ('hyunjin', '010-3333-9999')]
row_num = 2
for r, member in enumerate(members):
    for c, v in enumerate(member):
        sheet.cell(row = row_num, column = c + 1, value = v)
wb.save('/Users/yeryeongseo/Documents/repositories/pythonStudy/members.xlsx')
wb.close()