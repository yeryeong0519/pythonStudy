import openpyxl
wb = openpyxl.load_workbook('./sample.xlsx')
sheet = wb['mlist']
print(sheet.max_column, sheet.max_row)
print(sheet.cell(row1, column=1).value)
print(sheet.cell(row2, column=1).value)
wb.close()